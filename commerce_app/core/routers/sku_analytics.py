# commerce_app/core/routers/sku_analytics.py
from fastapi import APIRouter, HTTPException, Query, Depends
from commerce_app.core.db import get_conn
from commerce_app.auth.session_tokens import verify_shopify_session_token
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
from collections import defaultdict

router = APIRouter()


def get_shop_from_token(payload: Dict[str, Any] = Depends(verify_shopify_session_token)) -> str:
    """
    Extract shop domain from validated session token payload.
    The 'dest' claim contains the shop URL like: https://store.myshopify.com
    """
    # Extract from 'dest' field (standard Shopify session token format)
    dest = payload.get("dest", "")
    if dest:
        # Remove https:// prefix to match database format
        shop_domain = dest.replace("https://", "").replace("http://", "")
        # Also remove any trailing paths like /admin
        shop_domain = shop_domain.split("/")[0]
        return shop_domain
    
    # Fallback: extract from 'iss' field (format: https://store.myshopify.com/admin)
    iss = payload.get("iss", "")
    if iss:
        # Remove https:// and /admin
        shop = iss.replace("https://", "").replace("/admin", "").split("/")[0]
        return shop
    
    raise HTTPException(
        status_code=401,
        detail="Unable to extract shop domain from session token"
    )


@router.get("/sku-analytics/overview")
async def sku_overview(
    days: int = Query(default=30, ge=1, le=365, description="Number of days to analyze"),
    limit: int = Query(default=50, ge=1, le=200, description="Max SKUs to return"),
    sort_by: str = Query(default="revenue", regex="^(revenue|quantity|profit|margin)$", description="Sort field"),
    shop_domain: str = Depends(get_shop_from_token)
):
    """
    Get SKU-level performance overview.
    
    Returns:
    - SKU breakdown (quantity sold, revenue, profit, margin, AOV)
    - Sorted by specified metric
    - Includes COGS data when available
    """
    
    # First verify shop exists and get shop_id
    async with get_conn() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT shop_id FROM shopify.shops WHERE shop_domain = %s",
                (shop_domain,)
            )
            shop_row = await cur.fetchone()
            if not shop_row:
                raise HTTPException(404, "Shop not found")
            
            shop_id = shop_row[0]
            
            # Get order line items with product and COGS data
            await cur.execute(
                """
                SELECT 
                    oli.product_id,
                    oli.variant_id,
                    pv.sku,
                    p.title as line_item_name,
                    pv.title as variant_title,
                    oli.quantity,
                    oli.price,
                    p.title as product_title,
                    pv.cost as cogs_per_unit,
                    o.order_date
                FROM shopify.order_line_items oli
                INNER JOIN shopify.orders o ON oli.shop_id = o.shop_id AND oli.order_id = o.order_id
                LEFT JOIN shopify.products p ON oli.shop_id = p.shop_id AND oli.product_id = p.product_id
                LEFT JOIN shopify.product_variants pv ON oli.shop_id = pv.shop_id 
                    AND oli.product_id = pv.product_id 
                    AND oli.variant_id = pv.variant_id
                WHERE oli.shop_id = %s
                  AND o.order_date >= CURRENT_DATE - %s
                  AND o.financial_status IN ('paid', 'PAID', 'partially_paid','PARTIALLY_PAID')
                ORDER BY o.order_date DESC
                """,
                (shop_id, days)
            )
            
            line_items = await cur.fetchall()
    
    # Process SKUs to build performance metrics
    sku_stats = defaultdict(lambda: {
        "product_id": None,
        "variant_id": None,
        "sku": None,
        "product_title": None,
        "variant_title": None,
        "total_quantity": 0,
        "total_revenue": 0.0,
        "total_cost": 0.0,
        "total_profit": None,  # Will be None if no COGS data
        "has_cogs": False,
        "order_count": 0,
        "last_order_date": None
    })
    
    for line_item in line_items:
        (product_id, variant_id, sku, line_item_name, variant_title, 
         quantity, price, product_title, cogs_per_unit, created_at) = line_item
        
        # Create unique key for this SKU
        # Use variant_id if available, otherwise product_id
        if variant_id:
            key = f"v_{variant_id}"
        else:
            key = f"p_{product_id}"
        
        stats = sku_stats[key]
        
        # Set identifying information (first occurrence)
        if stats["product_id"] is None:
            stats["product_id"] = product_id
            stats["variant_id"] = variant_id
            stats["sku"] = sku
            stats["product_title"] = product_title or line_item_name
            stats["variant_title"] = variant_title
        
        # Aggregate metrics
        stats["total_quantity"] += quantity
        line_revenue = float(price) * quantity
        stats["total_revenue"] += line_revenue
        
        # Track COGS if available
        if cogs_per_unit is not None:
            stats["has_cogs"] = True
            line_cost = float(cogs_per_unit) * quantity
            stats["total_cost"] += line_cost
            
            # Initialize profit if this is first COGS entry
            if stats["total_profit"] is None:
                stats["total_profit"] = 0.0
            stats["total_profit"] += (line_revenue - line_cost)
        
        stats["order_count"] += 1
        
        # Track most recent order
        if stats["last_order_date"] is None or created_at > stats["last_order_date"]:
            stats["last_order_date"] = created_at
    
    # Format response
    skus = []
    for key, stats in sku_stats.items():
        # Calculate derived metrics
        avg_price = stats["total_revenue"] / stats["total_quantity"] if stats["total_quantity"] > 0 else 0
        cogs_per_unit = stats["total_cost"] / stats["total_quantity"] if stats["has_cogs"] and stats["total_quantity"] > 0 else None
        
        # Calculate profit margin if we have profit data
        profit_margin = None
        if stats["total_profit"] is not None and stats["total_revenue"] > 0:
            profit_margin = (stats["total_profit"] / stats["total_revenue"]) * 100
        
        skus.append({
            "product_id": stats["product_id"],
            "variant_id": stats["variant_id"],
            "sku": stats["sku"],
            "product_title": stats["product_title"],
            "variant_title": stats["variant_title"],
            "total_quantity": stats["total_quantity"],
            "total_revenue": round(stats["total_revenue"], 2),
            "total_profit": round(stats["total_profit"], 2) if stats["total_profit"] is not None else None,
            "avg_price": round(avg_price, 2),
            "cogs_per_unit": round(cogs_per_unit, 2) if cogs_per_unit is not None else None,
            "profit_margin": round(profit_margin, 2) if profit_margin is not None else None,
            "order_count": stats["order_count"],
            "last_order_date": stats["last_order_date"].isoformat() if stats["last_order_date"] else None,
            "has_cogs_data": stats["has_cogs"]
        })
    
    # Sort based on requested field
    sort_key_map = {
        "revenue": lambda x: x["total_revenue"],
        "quantity": lambda x: x["total_quantity"],
        "profit": lambda x: x["total_profit"] if x["total_profit"] is not None else -999999,
        "margin": lambda x: x["profit_margin"] if x["profit_margin"] is not None else -999999
    }
    
    skus.sort(key=sort_key_map.get(sort_by, lambda x: x["total_revenue"]), reverse=True)
    
    # Limit results
    limited_skus = skus[:limit]
    
    # Calculate summary statistics
    total_revenue = sum(s["total_revenue"] for s in skus)
    total_quantity = sum(s["total_quantity"] for s in skus)
    total_profit = sum(s["total_profit"] for s in skus if s["total_profit"] is not None)
    skus_with_profit = len([s for s in skus if s["total_profit"] is not None])
    
    return {
        "skus": limited_skus,
        "summary": {
            "total_skus": len(skus),
            "total_revenue": round(total_revenue, 2),
            "total_quantity": total_quantity,
            "total_profit": round(total_profit, 2) if skus_with_profit > 0 else None,
            "profit_data_available": skus_with_profit > 0,
            "skus_with_cogs": skus_with_profit,
            "skus_without_cogs": len(skus) - skus_with_profit
        },
        "date_range": {
            "start": (datetime.now() - timedelta(days=days)).date().isoformat(),
            "end": datetime.now().date().isoformat(),
            "days": days
        },
        "sort_by": sort_by,
        "limit": limit
    }


@router.get("/sku-analytics/trend")
async def sku_trend(
    days: int = Query(default=30, ge=1, le=90, description="Number of days to analyze"),
    group_by: str = Query(default="day", regex="^(day|week)$", description="Group by day or week"),
    top_n: int = Query(default=10, ge=1, le=20, description="Number of top SKUs to track"),
    shop_domain: str = Depends(get_shop_from_token)
):
    """
    Get SKU performance trend over time.
    
    Shows top N SKUs (by revenue) and their daily/weekly performance.
    """
    
    async with get_conn() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT shop_id FROM shopify.shops WHERE shop_domain = %s",
                (shop_domain,)
            )
            shop_row = await cur.fetchone()
            if not shop_row:
                raise HTTPException(404, "Shop not found")
            
            shop_id = shop_row[0]
            
            # First, get top N SKUs by revenue
            await cur.execute(
                """
                SELECT 
                    COALESCE(oli.variant_id::text, 'p_' || oli.product_id::text) as sku_key,
                    oli.product_id,
                    oli.variant_id,
                    pv.sku,
                    MAX(p.title) as product_title,
                    MAX(pv.title as variant_title) as variant_title,
                    SUM(oli.quantity * oli.price) as total_revenue
                FROM shopify.order_line_items oli
                INNER JOIN shopify.orders o ON oli.shop_id = o.shop_id AND oli.order_id = o.order_id
                LEFT JOIN shopify.products p ON oli.shop_id = p.shop_id AND oli.product_id = p.product_id
                WHERE oli.shop_id = %s
                  AND o.order_date >= CURRENT_DATE - %s
                  AND o.financial_status IN ('paid', 'PAID', 'partially_paid','PARTIALLY_PAID')
                GROUP BY sku_key, oli.product_id, oli.variant_id, pv.sku
                ORDER BY total_revenue DESC
                LIMIT %s
                """,
                (shop_id, days, top_n)
            )
            
            top_skus = await cur.fetchall()
            
            if not top_skus:
                return {
                    "series": [],
                    "group_by": group_by,
                    "date_range": {
                        "start": (datetime.now() - timedelta(days=days)).date().isoformat(),
                        "end": datetime.now().date().isoformat(),
                        "days": days
                    }
                }
            
            # Get the variant_ids and product_ids for filtering
            sku_filters = [(row[1], row[2]) for row in top_skus]  # (product_id, variant_id)
            
            # Determine date truncation
            date_trunc = "day" if group_by == "day" else "week"
            
            # Get time series data for these SKUs
            await cur.execute(
                """
                SELECT 
                    DATE_TRUNC(%s, o.order_date) as period,
                    oli.product_id,
                    oli.variant_id,
                    pv.sku,
                    MAX(p.title) as product_title,
                    MAX(pv.title as variant_title) as variant_title,
                    SUM(oli.quantity) as quantity,
                    SUM(oli.quantity * oli.price) as revenue
                FROM shopify.order_line_items oli
                INNER JOIN shopify.orders o ON oli.shop_id = o.shop_id AND oli.order_id = o.order_id
                LEFT JOIN shopify.products p ON oli.shop_id = p.shop_id AND oli.product_id = p.product_id
                WHERE oli.shop_id = %s
                  AND o.order_date >= CURRENT_DATE - %s
                  AND o.financial_status IN ('paid', 'PAID', 'partially_paid','PARTIALLY_PAID')
                GROUP BY period, oli.product_id, oli.variant_id, pv.sku
                ORDER BY period ASC
                """,
                (date_trunc, shop_id, days)
            )
            
            time_data = await cur.fetchall()
    
    # Build lookup for top SKUs
    top_sku_map = {}
    for row in top_skus:
        sku_key, product_id, variant_id, sku, product_title, variant_title, total_revenue = row
        
        # Create display name
        if variant_title:
            display_name = f"{product_title} - {variant_title}"
        else:
            display_name = product_title
        
        if variant_id:
            key = f"v_{variant_id}"
        else:
            key = f"p_{product_id}"
        
        top_sku_map[key] = {
            "product_id": product_id,
            "variant_id": variant_id,
            "sku": sku,
            "display_name": display_name,
            "total_revenue": float(total_revenue)
        }
    
    # Process time series data
    time_series = defaultdict(lambda: defaultdict(lambda: {"quantity": 0, "revenue": 0.0}))
    
    for row in time_data:
        period, product_id, variant_id, sku, product_title, variant_title, quantity, revenue = row
        
        # Create key
        if variant_id:
            key = f"v_{variant_id}"
        else:
            key = f"p_{product_id}"
        
        # Only include if in top N
        if key not in top_sku_map:
            continue
        
        # Format period as string
        period_str = period.strftime("%Y-%m-%d")
        
        time_series[key][period_str]["quantity"] += quantity
        time_series[key][period_str]["revenue"] += float(revenue or 0)
    
    # Format for frontend
    series = []
    for key, sku_info in top_sku_map.items():
        periods = time_series.get(key, {})
        
        data_points = [
            {
                "date": period,
                "quantity": stats["quantity"],
                "revenue": round(stats["revenue"], 2)
            }
            for period, stats in sorted(periods.items())
        ]
        
        series.append({
            "sku": sku_info["sku"],
            "display_name": sku_info["display_name"],
            "product_id": sku_info["product_id"],
            "variant_id": sku_info["variant_id"],
            "total_revenue": sku_info["total_revenue"],
            "data": data_points
        })
    
    # Sort series by total revenue
    series.sort(key=lambda x: x["total_revenue"], reverse=True)
    
    return {
        "series": series,
        "group_by": group_by,
        "date_range": {
            "start": (datetime.now() - timedelta(days=days)).date().isoformat(),
            "end": datetime.now().date().isoformat(),
            "days": days
        }
    }


@router.get("/sku-analytics/profit-leaders")
async def sku_profit_leaders(
    days: int = Query(default=30, ge=1, le=365, description="Number of days to analyze"),
    limit: int = Query(default=20, ge=1, le=50, description="Max SKUs to return"),
    shop_domain: str = Depends(get_shop_from_token)
):
    """
    Get top SKUs by profit margin and absolute profit.
    
    Only returns SKUs with COGS data available.
    """
    
    async with get_conn() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT shop_id FROM shopify.shops WHERE shop_domain = %s",
                (shop_domain,)
            )
            shop_row = await cur.fetchone()
            if not shop_row:
                raise HTTPException(404, "Shop not found")
            
            shop_id = shop_row[0]
            
            # Get SKUs with COGS data
            await cur.execute(
                """
                SELECT 
                    oli.product_id,
                    oli.variant_id,
                    pv.sku,
                    MAX(p.title) as product_title,
                    MAX(pv.title as variant_title) as variant_title,
                    SUM(oli.quantity) as total_quantity,
                    SUM(oli.quantity * oli.price) as total_revenue,
                    SUM(oli.quantity * pv.cost) as total_cost,
                    MAX(pv.cost) as cogs_per_unit
                FROM shopify.order_line_items oli
                INNER JOIN shopify.orders o ON oli.shop_id = o.shop_id AND oli.order_id = o.order_id
                INNER JOIN shopify.product_variants pv ON oli.shop_id = pv.shop_id 
                    AND oli.product_id = pv.product_id 
                    AND oli.variant_id = pv.variant_id
                LEFT JOIN shopify.products p ON oli.shop_id = p.shop_id AND oli.product_id = p.product_id
                WHERE oli.shop_id = %s
                  AND o.order_date >= CURRENT_DATE - %s
                  AND o.financial_status IN ('paid', 'PAID', 'partially_paid','PARTIALLY_PAID')
                GROUP BY oli.product_id, oli.variant_id, pv.sku
                HAVING SUM(oli.quantity * oli.price) > 0
                """,
                (shop_id, days)
            )
            
            rows = await cur.fetchall()
    
    if not rows:
        return {
            "top_by_margin": [],
            "top_by_profit": [],
            "message": "No SKUs with COGS data found in the selected period"
        }
    
    # Process and calculate metrics
    skus = []
    for row in rows:
        product_id, variant_id, sku, product_title, variant_title, total_quantity, total_revenue, total_cost, cogs_per_unit = row
        
        total_revenue = float(total_revenue or 0)
        total_cost = float(total_cost or 0)
        total_profit = total_revenue - total_cost
        profit_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
        
        # Create display name
        if variant_title:
            display_name = f"{product_title} - {variant_title}"
        else:
            display_name = product_title
        
        skus.append({
            "product_id": product_id,
            "variant_id": variant_id,
            "sku": sku,
            "display_name": display_name,
            "product_title": product_title,
            "variant_title": variant_title,
            "total_quantity": total_quantity,
            "total_revenue": round(total_revenue, 2),
            "total_cost": round(total_cost, 2),
            "total_profit": round(total_profit, 2),
            "profit_margin": round(profit_margin, 2),
            "cogs_per_unit": round(float(cogs_per_unit), 2)
        })
    
    # Sort by margin and get top N
    by_margin = sorted(skus, key=lambda x: x["profit_margin"], reverse=True)[:limit]
    
    # Sort by absolute profit and get top N
    by_profit = sorted(skus, key=lambda x: x["total_profit"], reverse=True)[:limit]
    
    return {
        "top_by_margin": by_margin,
        "top_by_profit": by_profit,
        "date_range": {
            "start": (datetime.now() - timedelta(days=days)).date().isoformat(),
            "end": datetime.now().date().isoformat(),
            "days": days
        }
    }


@router.get("/sku-analytics/export")
async def export_sku_analytics(
    days: int = Query(default=30, ge=1, le=365, description="Number of days to analyze"),
    limit: int = Query(default=50, ge=1, le=200, description="Max SKUs to return"),
    sort_by: str = Query(default="revenue", regex="^(revenue|quantity|profit|margin)$", description="Sort field"),
    shop_domain: str = Depends(get_shop_from_token)
):
    """
    Export SKU analytics to Excel file.
    
    Returns an Excel file with SKU performance data.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    # Reuse the same logic from sku_overview endpoint
    async with get_conn() as conn:
        async with conn.cursor() as cur:
            await cur.execute(
                "SELECT shop_id FROM shopify.shops WHERE shop_domain = %s",
                (shop_domain,)
            )
            shop_row = await cur.fetchone()
            if not shop_row:
                raise HTTPException(404, "Shop not found")
            
            shop_id = shop_row[0]
            
            # Get order line items with product and COGS data
            await cur.execute(
                """
                SELECT 
                    oli.product_id,
                    oli.variant_id,
                    pv.sku,
                    p.title as line_item_name,
                    pv.title as variant_title,
                    oli.quantity,
                    oli.price,
                    p.title as product_title,
                    pv.cost as cogs_per_unit,
                    o.order_date
                FROM shopify.order_line_items oli
                INNER JOIN shopify.orders o ON oli.shop_id = o.shop_id AND oli.order_id = o.order_id
                LEFT JOIN shopify.products p ON oli.shop_id = p.shop_id AND oli.product_id = p.product_id
                LEFT JOIN shopify.product_variants pv ON oli.shop_id = pv.shop_id 
                    AND oli.product_id = pv.product_id 
                    AND oli.variant_id = pv.variant_id
                WHERE oli.shop_id = %s
                  AND o.order_date >= CURRENT_DATE - %s
                  AND o.financial_status IN ('paid', 'PAID','partially_paid','PARTIALLY_PAID')
                ORDER BY o.order_date DESC
                """,
                (shop_id, days)
            )
            
            line_items = await cur.fetchall()
    
    # Process SKUs (same logic as overview endpoint)
    sku_stats = defaultdict(lambda: {
        "product_id": None,
        "variant_id": None,
        "sku": None,
        "product_title": None,
        "variant_title": None,
        "total_quantity": 0,
        "total_revenue": 0.0,
        "total_cost": 0.0,
        "total_profit": None,
        "has_cogs": False,
        "order_count": 0,
        "last_order_date": None
    })
    
    for line_item in line_items:
        (product_id, variant_id, sku, line_item_name, variant_title, 
         quantity, price, product_title, cogs_per_unit, created_at) = line_item
        
        if variant_id:
            key = f"v_{variant_id}"
        else:
            key = f"p_{product_id}"
        
        stats = sku_stats[key]
        
        if stats["product_id"] is None:
            stats["product_id"] = product_id
            stats["variant_id"] = variant_id
            stats["sku"] = sku
            stats["product_title"] = product_title or line_item_name
            stats["variant_title"] = variant_title
        
        stats["total_quantity"] += quantity
        line_revenue = float(price) * quantity
        stats["total_revenue"] += line_revenue
        
        if cogs_per_unit is not None:
            stats["has_cogs"] = True
            line_cost = float(cogs_per_unit) * quantity
            stats["total_cost"] += line_cost
            
            if stats["total_profit"] is None:
                stats["total_profit"] = 0.0
            stats["total_profit"] += (line_revenue - line_cost)
        
        stats["order_count"] += 1
        
        if stats["last_order_date"] is None or created_at > stats["last_order_date"]:
            stats["last_order_date"] = created_at
    
    # Format and sort SKUs
    skus = []
    for key, stats in sku_stats.items():
        avg_price = stats["total_revenue"] / stats["total_quantity"] if stats["total_quantity"] > 0 else 0
        cogs_per_unit = stats["total_cost"] / stats["total_quantity"] if stats["has_cogs"] and stats["total_quantity"] > 0 else None
        
        profit_margin = None
        if stats["total_profit"] is not None and stats["total_revenue"] > 0:
            profit_margin = (stats["total_profit"] / stats["total_revenue"]) * 100
        
        skus.append({
            "product_id": stats["product_id"],
            "variant_id": stats["variant_id"],
            "sku": stats["sku"],
            "product_title": stats["product_title"],
            "variant_title": stats["variant_title"],
            "total_quantity": stats["total_quantity"],
            "total_revenue": stats["total_revenue"],
            "total_profit": stats["total_profit"],
            "avg_price": avg_price,
            "cogs_per_unit": cogs_per_unit,
            "profit_margin": profit_margin,
            "order_count": stats["order_count"],
            "last_order_date": stats["last_order_date"],
            "has_cogs_data": stats["has_cogs"]
        })
    
    # Sort
    sort_key_map = {
        "revenue": lambda x: x["total_revenue"],
        "quantity": lambda x: x["total_quantity"],
        "profit": lambda x: x["total_profit"] if x["total_profit"] is not None else -999999,
        "margin": lambda x: x["profit_margin"] if x["profit_margin"] is not None else -999999
    }
    
    skus.sort(key=sort_key_map.get(sort_by, lambda x: x["total_revenue"]), reverse=True)
    limited_skus = skus[:limit]
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "SKU Analytics"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = [
        "Product Title",
        "Variant",
        "SKU",
        "Quantity Sold",
        "Total Revenue",
        "Total Profit",
        "Profit Margin %",
        "Avg Price",
        "COGS/Unit",
        "Order Count",
        "Last Order Date"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_num, sku in enumerate(limited_skus, 2):
        ws.cell(row=row_num, column=1, value=sku["product_title"])
        ws.cell(row=row_num, column=2, value=sku["variant_title"] or "")
        ws.cell(row=row_num, column=3, value=sku["sku"] or "")
        ws.cell(row=row_num, column=4, value=sku["total_quantity"])
        ws.cell(row=row_num, column=5, value=round(sku["total_revenue"], 2))
        ws.cell(row=row_num, column=6, value=round(sku["total_profit"], 2) if sku["total_profit"] is not None else "No COGS")
        ws.cell(row=row_num, column=7, value=round(sku["profit_margin"], 2) if sku["profit_margin"] is not None else "N/A")
        ws.cell(row=row_num, column=8, value=round(sku["avg_price"], 2))
        ws.cell(row=row_num, column=9, value=round(sku["cogs_per_unit"], 2) if sku["cogs_per_unit"] is not None else "Not set")
        ws.cell(row=row_num, column=10, value=sku["order_count"])
        ws.cell(row=row_num, column=11, value=sku["last_order_date"].strftime("%Y-%m-%d") if sku["last_order_date"] else "Never")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as streaming response
    filename = f"sku_analytics_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )