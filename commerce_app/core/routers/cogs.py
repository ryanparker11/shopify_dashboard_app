from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from commerce_app.core.db import get_conn
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import io
from typing import List, Dict, Any
from datetime import datetime

router = APIRouter()

@router.get("/cogs/download-template")
async def download_cogs_template(shop_domain: str):
    """
    Generate and download COGS upload template with VARIANT_ID, SKU, NAME, and VARIANT pre-filled.
    User only needs to fill in the COGS column.
    Uses variant_id as the primary key for matching.
    """
    try:
        # Fetch products and variants from database
        sql = """
        SELECT 
            pv.variant_id,
            pv.sku,
            p.title as product_name,
            pv.title as variant_title
        FROM shopify.product_variants pv
        JOIN shopify.products p ON pv.product_id = p.product_id
        WHERE p.shop_id = (SELECT shop_id FROM shopify.shops WHERE shop_domain = %s)
        ORDER BY p.title, pv.title;
        """
        
        async with get_conn() as conn:
            async with conn.cursor() as cur:
                await cur.execute(sql, (shop_domain,))
                products_data = await cur.fetchall()
                
                if not products_data:
                    raise HTTPException(404, "No products found for this shop")
        
        # Generate Excel template
        wb = Workbook()
        sheet = wb.active
        sheet.title = "COGS Upload"
        
        # Header row styling
        headers = ['VARIANT_ID', 'SKU', 'NAME', 'VARIANT', 'COGS']
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', start_color='4472C4')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        sheet.column_dimensions['A'].width = 15  # VARIANT_ID
        sheet.column_dimensions['B'].width = 20  # SKU
        sheet.column_dimensions['C'].width = 40  # NAME
        sheet.column_dimensions['D'].width = 30  # VARIANT
        sheet.column_dimensions['E'].width = 15  # COGS
        
        # Populate data rows
        for row_idx, (variant_id, sku, product_name, variant_title) in enumerate(products_data, start=2):
            sheet.cell(row=row_idx, column=1, value=variant_id)
            sheet.cell(row=row_idx, column=2, value=sku or '')
            sheet.cell(row=row_idx, column=3, value=product_name or '')
            sheet.cell(row=row_idx, column=4, value=variant_title if variant_title != 'Default Title' else 'Default')
            
            # COGS column - yellow highlight for user input
            cogs_cell = sheet.cell(row=row_idx, column=5)
            cogs_cell.fill = PatternFill('solid', start_color='FFFF00')
        
        # Freeze header row
        sheet.freeze_panes = 'A2'
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as downloadable file
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename=cogs_upload_template.xlsx'
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating template: {str(e)}")


@router.get("/cogs/summary")
async def cogs_summary(shop_domain: str):
    """
    Get COGS summary statistics for the shop
    """
    sql = """
    SELECT
        COUNT(DISTINCT pv.variant_id)::int AS total_variants,
        COUNT(DISTINCT CASE WHEN pv.cost IS NOT NULL THEN pv.variant_id END)::int AS variants_with_cogs,
        COALESCE(AVG(pv.cost), 0)::numeric AS avg_cogs
    FROM shopify.product_variants pv
    JOIN shopify.products p ON pv.product_id = p.product_id
    WHERE p.shop_id = (SELECT shop_id FROM shopify.shops WHERE shop_domain = %s);
    """
    
    async with get_conn() as conn:
        async with conn.cursor() as cur:
            await cur.execute(sql, (shop_domain,))
            row = await cur.fetchone()
            if row is None:
                raise HTTPException(404, "Shop not found")
            
            total_variants, variants_with_cogs, avg_cogs = row
            
            return {
                "total_variants": total_variants,
                "variants_with_cogs": variants_with_cogs,
                "variants_without_cogs": total_variants - variants_with_cogs,
                "avg_cogs": float(avg_cogs) if avg_cogs else 0.0,
                "cogs_coverage_percentage": round((variants_with_cogs / total_variants * 100), 2) if total_variants > 0 else 0.0
            }


@router.post("/cogs/upload-template")
async def upload_cogs_template(shop_domain: str, file: UploadFile = File(...)):
    """
    Upload completed COGS template and update product costs.
    Uses VARIANT_ID as the primary matching key (more reliable than SKU).
    This will overwrite existing COGS data.
    """
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(400, "File must be an Excel file (.xlsx or .xls)")
        
        # Read the uploaded file
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Validate required columns
        required_columns = ['VARIANT_ID', 'COGS']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(400, f"Missing required columns: {', '.join(missing_columns)}")
        
        # Clean the data
        df = df.dropna(subset=['VARIANT_ID', 'COGS'])  # Remove rows without VARIANT_ID or COGS
        df['VARIANT_ID'] = pd.to_numeric(df['VARIANT_ID'], errors='coerce', downcast='integer')
        df['COGS'] = pd.to_numeric(df['COGS'], errors='coerce')
        df = df.dropna(subset=['VARIANT_ID', 'COGS'])  # Remove rows where conversion failed
        
        if len(df) == 0:
            raise HTTPException(400, "No valid COGS data found in the file")
        
        # Get shop_id
        async with get_conn() as conn:
            async with conn.cursor() as cur:
                await cur.execute(
                    "SELECT shop_id FROM shopify.shops WHERE shop_domain = %s",
                    (shop_domain,)
                )
                shop_row = await cur.fetchone()
                if not shop_row:
                    raise HTTPException(404, "Shop not found")
                shop_id = shop_row[0]
        
        # Update COGS for each variant_id
        updated_count = 0
        skipped_count = 0
        errors = []
        
        async with get_conn() as conn:
            async with conn.cursor() as cur:
                for _, row in df.iterrows():
                    variant_id = int(row['VARIANT_ID'])
                    cogs = float(row['COGS'])
                    
                    try:
                        # Update product variant cost using variant_id
                        update_sql = """
                        UPDATE shopify.product_variants pv
                        SET cost = %s,
                            updated_at = NOW()
                        FROM shopify.products p
                        WHERE pv.product_id = p.product_id
                          AND p.shop_id = %s
                          AND pv.variant_id = %s
                        """
                        await cur.execute(update_sql, (cogs, shop_id, variant_id))
                        
                        if cur.rowcount > 0:
                            updated_count += 1
                        else:
                            skipped_count += 1
                            errors.append(f"Variant ID not found: {variant_id}")
                    
                    except Exception as e:
                        skipped_count += 1
                        errors.append(f"Error updating Variant ID {variant_id}: {str(e)}")
                
                # Commit the transaction
                await conn.commit()
        
        return {
            "success": True,
            "updated_count": updated_count,
            "skipped_count": skipped_count,
            "total_rows": len(df),
            "errors": errors[:10] if errors else None,  # Return first 10 errors
            "message": f"Successfully updated {updated_count} products"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@router.get("/cogs/profit-analysis")
async def profit_analysis(shop_domain: str, days: int = 30):
    """
    Calculate profit metrics based on orders and COGS data.
    Returns revenue, COGS, profit, and margin for the specified period.
    Filters out line items with NULL variant_id to ensure accurate COGS matching.
    """
    try:
        sql = """
        WITH order_items AS (
            SELECT 
                li.order_id,
                li.product_id,
                li.variant_id,
                li.quantity,
                li.price,
                (li.quantity * li.price) as line_revenue,
                pv.cost as unit_cogs,
                (li.quantity * COALESCE(pv.cost, 0)) as line_cogs
            FROM shopify.order_line_items li
            JOIN shopify.orders o ON li.order_id = o.order_id
            LEFT JOIN shopify.product_variants pv ON li.variant_id = pv.variant_id
            WHERE o.shop_id = (SELECT shop_id FROM shopify.shops WHERE shop_domain = %s)
              AND o.created_at >= NOW() - INTERVAL '1 day' * %s
              AND o.financial_status IN ('paid', 'partially_paid')
              AND li.variant_id IS NOT NULL
        )
        SELECT 
            COALESCE(SUM(line_revenue), 0)::numeric as total_revenue,
            COALESCE(SUM(line_cogs), 0)::numeric as total_cogs,
            COALESCE(SUM(line_revenue) - SUM(line_cogs), 0)::numeric as gross_profit,
            CASE 
                WHEN SUM(line_revenue) > 0 
                THEN ROUND((SUM(line_revenue) - SUM(line_cogs)) / SUM(line_revenue) * 100, 2)
                ELSE 0 
            END as profit_margin_pct,
            COUNT(DISTINCT order_id)::int as order_count,
            COUNT(DISTINCT CASE WHEN unit_cogs IS NULL THEN variant_id END)::int as items_without_cogs
        FROM order_items;
        """
        
        async with get_conn() as conn:
            async with conn.cursor() as cur:
                await cur.execute(sql, (shop_domain, days))
                row = await cur.fetchone()
                
                if not row:
                    raise HTTPException(404, "No data found")
                
                total_revenue, total_cogs, gross_profit, profit_margin_pct, order_count, items_without_cogs = row
                
                return {
                    "period_days": days,
                    "total_revenue": float(total_revenue),
                    "total_cogs": float(total_cogs),
                    "gross_profit": float(gross_profit),
                    "profit_margin_percentage": float(profit_margin_pct),
                    "order_count": order_count,
                    "items_without_cogs": items_without_cogs,
                    "has_complete_data": items_without_cogs == 0
                }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating profit: {str(e)}")


@router.get("/cogs/profit-by-product")
async def profit_by_product(shop_domain: str, days: int = 30, limit: int = 10):
    """
    Get top products by profit with COGS breakdown.
    Filters out line items with NULL variant_id to ensure accurate COGS matching.
    """
    try:
        sql = """
        SELECT 
            p.title as product_name,
            SUM(li.quantity)::int as units_sold,
            COALESCE(SUM(li.quantity * li.price), 0)::numeric as revenue,
            COALESCE(SUM(li.quantity * pv.cost), 0)::numeric as cogs,
            COALESCE(SUM(li.quantity * li.price) - SUM(li.quantity * pv.cost), 0)::numeric as profit,
            CASE 
                WHEN SUM(li.quantity * li.price) > 0 
                THEN ROUND((SUM(li.quantity * li.price) - SUM(li.quantity * pv.cost)) / SUM(li.quantity * li.price) * 100, 2)
                ELSE 0 
            END as margin_pct
        FROM shopify.order_line_items li
        JOIN shopify.orders o ON li.order_id = o.order_id
        JOIN shopify.products p ON li.product_id = p.product_id
        LEFT JOIN shopify.product_variants pv ON li.variant_id = pv.variant_id
        WHERE o.shop_id = (SELECT shop_id FROM shopify.shops WHERE shop_domain = %s)
          AND o.created_at >= NOW() - INTERVAL '1 day' * %s
          AND o.financial_status IN ('paid', 'partially_paid')
          AND li.variant_id IS NOT NULL
        GROUP BY p.product_id, p.title
        ORDER BY profit DESC
        LIMIT %s;
        """
        
        async with get_conn() as conn:
            async with conn.cursor() as cur:
                await cur.execute(sql, (shop_domain, days, limit))
                rows = await cur.fetchall()
                
                products = []
                for row in rows:
                    product_name, units_sold, revenue, cogs, profit, margin_pct = row
                    products.append({
                        "product_name": product_name,
                        "units_sold": units_sold,
                        "revenue": float(revenue),
                        "cogs": float(cogs),
                        "profit": float(profit),
                        "margin_percentage": float(margin_pct)
                    })
                
                return {
                    "products": products,
                    "period_days": days
                }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching product profit: {str(e)}")